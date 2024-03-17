import openpyxl

class SolarMeta:
    def __init__(self, input_file_path='input_data.xlsx', sheet_name='solar'):
        wb = openpyxl.load_workbook(input_file_path, data_only=True)
        sheet = wb[sheet_name]

        # Set attributes from the Excel sheet
        self.capital_cost_baseline = sheet['A2'].value
        self.fixed_opex_baseline = sheet['B2'].value
        self.useful_life = sheet['C2'].value
        self.opex_inflation_rate = sheet['D2'].value
        self.specific_yield = sheet['E2'].value
        self.mw_space_required = sheet['F2'].value
        self.annual_energy_per_mw = sheet['G2'].value
        self.num_failures_year = sheet['H2'].value
        self.avg_failure_time = sheet['I2'].value
        self.degradation = sheet['J2'].value
        self.existing_cap_cost = sheet['K2'].value
        self.output_data = {}

        # Monthly data
        for row in range(6, 18):  # Rows 6 to 17 for months 1 to 12
            month = row - 5
            self.output_data[month] = {
                'mnth_ener_op_per_MW': sheet.cell(row=row, column=2).value,  # Column B
                'output_proportion_of_year': sheet.cell(row=row, column=3).value,  # Column C
                'average_sun_available_hours': sheet.cell(row=row, column=4).value,  # Column D
                'average_sun_unavailable_hours': sheet.cell(row=row, column=5).value  # Column E
            }

        # Hourly data for each month
        current_row = 22
        for month in range(1, 13):  # For each month from 1 to 12
            for hour in range(1, 25):  # For each hour from 1 to 24
                self.output_data[month][hour] = {
                    'kw_pv_output_per_mw': sheet.cell(row=current_row, column=3).value  # Column C
                }
                current_row += 1  # Move to the next row

class WindMeta:
    def __init__(self, input_file_path='input_data.xlsx', sheet_name='wind'):
        wb = openpyxl.load_workbook(input_file_path, data_only=True)
        sheet = wb[sheet_name]

        # Set attributes from the Excel sheet
        self.capital_cost_baseline = sheet['A2'].value
        self.fixed_opex_baseline = sheet['B2'].value
        self.useful_life = sheet['C2'].value
        self.opex_inflation_rate = sheet['D2'].value
        self.days_of_year = sheet['E2'].value
        self.hours_of_day = sheet['F2'].value
        self.power_density_at_site = sheet['G2'].value
        self.wind_class_applicable = sheet['H2'].value
        self.selected_turbine_rotor_diameter = sheet['I2'].value
        self.swept_area = sheet['J2'].value
        self.tower_height = sheet['K2'].value
        self.power_density_of_turbine = sheet['L2'].value
        self.avg_speed = sheet['M2'].value
        self.turbine_rated_power = sheet['N2'].value
        self.output_multiplier_1 = sheet['O2'].value
        self.output_multiplier_2 = sheet['P2'].value
        self.output_multiplier_3 = sheet['Q2'].value
        self.space_required_for_2mw = sheet['R2'].value
        self.num_failures_year = sheet['S2'].value
        self.avg_failure_time = sheet['T2'].value
        self.degradation = sheet['U2'].value
        self.existing_cap_cost = 0

        self.output_data = {}

        # Monthly data
        for row in range(6, 18):  # Rows 6 to 17 for months 1 to 12
            month = row - 5
            self.output_data[month] = {
                'per_month_energy_output': sheet.cell(row=row, column=2).value,  # Column B
                'output_prop': sheet.cell(row=row, column=3).value  # Column C
            }

        # Hourly data for each month
        current_row = 21
        for month in range(1, 13):  # For each month from 1 to 12
            for hour in range(1, 25):  # For each hour from 1 to 24
                self.output_data[month][hour] = {
                    'windspeed_at_100m': sheet.cell(row=current_row, column=3).value  # Column C
                }
                current_row += 1  # Move to the next row

class PPAMeta:
    def __init__(self, input_file_path='input_data.xlsx', sheet_name='ppa'):
        workbook = openpyxl.load_workbook(input_file_path, data_only=True)
        sheet = workbook[sheet_name]

        self.capital_cost_baseline = sheet['C3'].value
        self.existing_cap_cost = sheet['C4'].value
        self.tariff_baseline_fixed = sheet['C5'].value
        self.tariff_baseline_var = sheet['C6'].value
        self.num_failures_year = sheet['C7'].value
        self.avg_failure_time = sheet['C8'].value
        self.opex_inflation_rate = sheet['C9'].value
        self.time_take_load_prim = sheet['C10'].value
        self.co2_emission = sheet['C11'].value
        self.output_data = {}

        sheet = workbook['solar']
        # Monthly data
        for row in range(6, 18):  # Rows 6 to 17 for months 1 to 12
            month = row - 5
            self.output_data[month] = {
                'mnth_ener_op_per_MW': sheet.cell(row=row, column=2).value,  # Column B
                'output_proportion_of_year': sheet.cell(row=row, column=3).value,  # Column C
                'average_sun_available_hours': sheet.cell(row=row, column=4).value,  # Column D
                'average_sun_unavailable_hours': sheet.cell(row=row, column=5).value  # Column E
            }

        # Hourly data for each month
        current_row = 22
        for month in range(1, 13):  # For each month from 1 to 12
            for hour in range(1, 25):  # For each hour from 1 to 24
                self.output_data[month][hour] = {
                    'kw_pv_output_per_mw': sheet.cell(row=current_row, column=3).value  # Column C
                }
                current_row += 1  # Move to the next row

class GridMeta:
    def __init__(self, input_file_path='input_data.xlsx', sheet_name='grid'):
        workbook = openpyxl.load_workbook(input_file_path, data_only=True)
        sheet = workbook[sheet_name]

        self.capital_cost_baseline = sheet['A2'].value
        self.existing_cap_cost = sheet['B2'].value
        self.tariff_baseline_fixed = sheet['C2'].value
        self.tariff_baseline_var_offpeak = sheet['D2'].value
        self.tariff_baseline_var_peak = sheet['E2'].value
        self.num_failures_year = sheet['F2'].value
        self.avg_failure_time = sheet['G2'].value
        self.opex_inflation_rate = sheet['H2'].value
        self.time_take_load_prim = sheet['I2'].value
        self.co2_emission = sheet['J2'].value


class TriFuelGenMeta:

    def __init__(self, input_file_path='input_data.xlsx', sheet_name='thermal_sources'):

        wb = openpyxl.load_workbook(input_file_path, data_only=True)
        sheet = wb[sheet_name]

        self.operating_baseline = sheet['B4'].value
        self.capital_cost_baseline = sheet['B5'].value
        self.fixed_opex_baseline = sheet['B6'].value
        self.var_opex_baseline = sheet['B7'].value
        self.useful_life = sheet['B8'].value
        self.current_running_hours = sheet['B9'].value
        self.opex_inflation_rate = sheet['B10'].value
        self.depreciation_rate = sheet['B11'].value
        self.cooling_load_feeding_capability = sheet['B12'].value
        self.min_loading = sheet['B13'].value
        self.max_loading = sheet['B14'].value
        self.num_failures_year = sheet['B15'].value
        self.avg_failure_time = sheet['B16'].value
        self.time_take_load_prim = sheet['B17'].value
        self.time_take_load_backup = sheet['B18'].value
        self.co2_emission = sheet['B19'].value
        self.co2_emission_sec = sheet['B20'].value
        self.gas_op_prop = sheet['B21'].value
        self.degradation = sheet['B22'].value
        self.existing_cap_cost = 0

class GasGenMeta():

     def __init__(self, input_file_path='input_data.xlsx', sheet_name='thermal_sources'):

        wb = openpyxl.load_workbook(input_file_path, data_only=True)
        sheet = wb[sheet_name]

        self.operating_baseline = sheet['F4'].value
        self.capital_cost_baseline = sheet['F5'].value
        self.fixed_opex_baseline = sheet['F6'].value
        self.var_opex_baseline = sheet['F7'].value
        self.useful_life = sheet['F8'].value
        self.current_running_hours = sheet['F9'].value
        self.opex_inflation_rate = sheet['F10'].value
        self.depreciation_rate = sheet['F11'].value
        self.cooling_load_feeding_capability = sheet['F12'].value
        self.min_loading = sheet['F13'].value
        self.max_loading = sheet['F14'].value
        self.num_failures_year = sheet['F15'].value
        self.avg_failure_time = sheet['F16'].value
        self.time_take_load_prim = sheet['F17'].value
        self.time_take_load_backup = sheet['F18'].value
        self.co2_emission = sheet['F19'].value
        self.degradation = sheet['F20'].value
        self.existing_cap_cost = sheet['F21'].value

class ExistingGasGenMeta():

     def __init__(self, input_file_path='input_data.xlsx', sheet_name='thermal_sources'):

        wb = openpyxl.load_workbook(input_file_path, data_only=True)
        sheet = wb[sheet_name]

        self.operating_baseline = sheet['R4'].value
        self.capital_cost_baseline = sheet['R5'].value
        self.fixed_opex_baseline = sheet['R6'].value
        self.var_opex_baseline = sheet['R7'].value
        self.useful_life = sheet['R8'].value
        self.current_running_hours = sheet['R9'].value
        self.opex_inflation_rate = sheet['R10'].value
        self.depreciation_rate = sheet['R11'].value
        self.cooling_load_feeding_capability = sheet['R12'].value
        self.min_loading = sheet['R13'].value
        self.max_loading = sheet['R14'].value
        self.num_failures_year = sheet['R15'].value
        self.avg_failure_time = sheet['R16'].value
        self.time_take_load_prim = sheet['R17'].value
        self.time_take_load_backup = sheet['R18'].value
        self.co2_emission = sheet['R19'].value
        self.degradation = sheet['R20'].value
        self.existing_cap_cost = sheet['R21'].value
        self.ingestion_cost = sheet['R22'].value

class HFOGenMeta():

    def __init__(self, input_file_path='input_data.xlsx', sheet_name='thermal_sources'):

        wb = openpyxl.load_workbook(input_file_path, data_only=True)
        sheet = wb[sheet_name]

        self.operating_baseline = sheet['J4'].value
        self.capital_cost_baseline = sheet['J5'].value
        self.fixed_opex_baseline = sheet['J6'].value
        self.var_opex_baseline = sheet['J7'].value
        self.useful_life = sheet['J8'].value
        self.current_running_hours = sheet['J9'].value
        self.opex_inflation_rate = sheet['J10'].value
        self.depreciation_rate = sheet['J11'].value
        self.cooling_load_feeding_capability = sheet['J12'].value
        self.min_loading = sheet['J13'].value
        self.max_loading = sheet['J14'].value
        self.num_failures_year = sheet['J15'].value
        self.avg_failure_time = sheet['J16'].value
        self.time_take_load_prim = sheet['J17'].value
        self.time_take_load_backup = sheet['J18'].value
        self.co2_emission = sheet['J19'].value
        self.degradation = sheet['J20'].value
        self.existing_cap_cost = 0

class DGenMeta():

    def __init__(self, input_file_path='input_data.xlsx', sheet_name='thermal_sources'):

        wb = openpyxl.load_workbook(input_file_path, data_only=True)
        sheet = wb[sheet_name]

        self.operating_baseline = sheet['N4'].value
        self.capital_cost_baseline = sheet['N5'].value
        self.fixed_opex_baseline = sheet['N6'].value
        self.var_opex_baseline = sheet['N7'].value
        self.useful_life = sheet['N8'].value
        self.current_running_hours = sheet['N9'].value
        self.opex_inflation_rate = sheet['N10'].value
        self.depreciation_rate = sheet['N11'].value
        self.min_loading = sheet['N13'].value
        self.max_loading = sheet['N14'].value
        self.num_failures_year = 0
        self.avg_failure_time = 0
        self.time_take_load_prim = sheet['N17'].value
        self.co2_emission = sheet['N18'].value
        self.degradation = sheet['N19'].value
        self.existing_cap_cost = 0

class BESSMeta():

    def __init__(self, input_file_path='input_data.xlsx', sheet_name='solar'):

        wb = openpyxl.load_workbook(input_file_path, data_only=True)
        sheet = wb[sheet_name]

        self.capital_cost_baseline = sheet['I7'].value
        self.fixed_opex_baseline = sheet['I8'].value
        self.useful_life = sheet['I9'].value
        self.depr_rate = sheet['I10'].value
        self.opex_inflation_rate = sheet['I11'].value
        self.existing_cap_cost = 0


class FuelTariff:

    def __init__(self, input_file_path='input_data.xlsx', sheet_name='tariff'):

        wb = openpyxl.load_workbook(input_file_path, data_only=True)
        sheet = wb[sheet_name]

        self.tariffs = {
            'NG': {
                'tariff': sheet['B5'].value,
                'inflation': sheet['B6'].value,
                'co2_emission': sheet['B7'].value
            },
            'RLNG': {
                'tariff': sheet['B10'].value,
                'inflation': sheet['B11'].value,
                'co2_emission': sheet['B12'].value
            },
            'LPG': {
                'tariff': sheet['B15'].value,
                'inflation': sheet['B16'].value,
                'co2_emission': sheet['B17'].value
            },
            'Biogas': {
                'tariff': sheet['B20'].value,
                'inflation': sheet['B21'].value,
                'co2_emission': sheet['B22'].value
            },
            'HFO': {
                'tariff': sheet['B25'].value,
                'inflation': sheet['B26'].value,
                'co2_emission': sheet['B27'].value
            },
            'Diesel': {
                'tariff': sheet['B30'].value,
                'inflation': sheet['B31'].value,
                'co2_emission': sheet['B32'].value
            }
        }

    def get_tariff_and_inflation(self, fuel_name):
        return self.tariffs.get(fuel_name, {'tariff': None, 'inflation': None, 'co2_emission': None})


