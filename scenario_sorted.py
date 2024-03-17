import datetime
import openpyxl
import pandas as pd
import math
import random

from source_meta import FuelTariff

class Scenario:
    def __init__(self, name, client_name, input_file_path='input_data.xlsx', n=5):
        self.name = name
        self.client_name = client_name
        self.timestamp = datetime.datetime.now()
        self.scenario_spec = {}
        self.sources_dict = {}
        self.sources_list = []
        self.ip_site_data = {}
        self.ip_load_data = {}
        self.ip_enr_data = {}
        #OUTPUT DATAFRAMES
        self.power_df = []
        self.energy_df = []
        self.capex_df = []
        self.opex_df = []
        self.emissions_df = []
        # SUMMARY OUTPUT DATAFRAMES
        self.summary_df = None
        self.power_fulfillment = None
        self.unserved_power = None
        self.energy_fulfillment = None
        self.unserved_energy = None
        self.power_summary_df = None
        self.energy_summary_df = None
        self.energy_summary_concise_df = None
        self.emissions_summary_df = None
        self.opex_summary_df = None
        self.opex_summary_concise_df = None

        self.n = n
        self.load_from_excel(input_file_path, 'site_load')

    def generate_results(self):

        self.get_scenario_config()
        self.power_calculation2()
        self.energy_calculation()
        self.emissions_calculation()
        self.capex_calculation2()
        self.opex_calculation()

    def generate_summaries(self):
        #Summmaries
        self.gen_pwr_fulfillment()
        self.gen_enr_fulfillment()
        self.gen_annual_pwr_summary()
        self.gen_annual_enr_summary()
        self.gen_annual_enr_summary_concise()
        self.gen_annual_emissions_summary()
        self.gen_annual_opex_summary()
        self.gen_annual_opex_summary_concise()
        self.gen_annual_summary()

    def add_source(self, source):

        self.sources_dict[source.source_type] = source
        self.sources_list.append(source)
        self.sources_list.sort(key= lambda src: src.priority)

    @staticmethod
    def available_sources(all_sources=True):
        if all_sources:
            return ["Solar", "Wind", "Gas Generator", "Existing Gas Generators","HFO Generator",
                    "HFO+Gas Generator", "PPA","Grid","Diesel Generator", "BESS"]
        else:
            return ["Solar", "Wind", "Gas Generator", "Existing Gas Generators","HFO Generator",
                    "HFO+Gas Generator", "PPA", "Grid","Diesel Generator"]

    @staticmethod
    def stable_sources(include_backup):
        
        if include_backup:

            return ["Gas Generator", "Existing Gas Generators","HFO Generator", "HFO+Gas Generator", 
                    "Grid", "Diesel Generator"]
        else:
            return ["Gas Generator", "Existing Gas Generators","HFO Generator", "HFO+Gas Generator"
                    ,"Grid"]

    @staticmethod
    def available_gas_types():
        return ["Natural Gas", "RLNG", "LPG", "Bio-gas"]

    @staticmethod
    def derating_factor(fuel_type):
        # Dictionary mapping fuel types to their respective derating factors
        derating_factors = {
            "Natural Gas": 1.0,
            "RLNG": 1.0,
            "NG": 1.0,
            "LPG": 1.0,
            "Bio-gas": 1.0
        }

        # Return the derating factor if found, otherwise return 1.0
        return derating_factors.get(fuel_type, 1.0)

    def load_from_excel(self, input_file_path, sheet_name):
        try:
            # Load site data
            wb = openpyxl.load_workbook(input_file_path, data_only=True)
            sheet = wb[sheet_name]

            site_keys = [
                "cop_of_electric_chiller",
                "cooling_hours_per_day",
                "critical_prod_load_prop",
                "summ_day_cool_dem_prop",
                "summ_nht_cool_dem_prop",
                "wint_day_cool_dem_prop",
                "wint_nht_cool_dem_prop",
                "land_area_for_renewables",
                "rooftop_area_for_solar",
                "capital_inflation_rate",
                "fail_loss_immediate",
                "fail_loss_over_time"
            ]

            for index, key in enumerate(site_keys):
                self.ip_site_data[key] = sheet.cell(row=index + 5, column=3).value

            # Load load details
            year_keys = [i for i in range(1, 11)]
            load_data_keys = [
                "max_dem_load_day",
                "max_dem_load_night",
                "running_load_prop",
                "cool_req_in_tr",
                "cool_elect_load",
                "crit_load_prop"
            ]

            for y_index, year_key in enumerate(year_keys):
                year_data = {}
                for l_index, load_key in enumerate(load_data_keys):
                    year_data[load_key] = sheet.cell(row=y_index + 4, column=l_index + 9).value
                self.ip_load_data[year_key] = year_data

            month_keys = [i for i in range(1, 13)]
            energy_data_keys = [
                "grid_enr_pk_to_nonpk_prop",
                "prod_enr_consump_per_MW_MD",
                "length_day",
                "length_night",
                "season",
                "days"
            ]

            for m_index, m_key in enumerate(month_keys):
                month_data = {}
                for e_index, energy_key in enumerate(energy_data_keys):
                    month_data[energy_key] = sheet.cell(row=m_index + 20, column=e_index + 9).value
                self.ip_enr_data[m_key] = month_data

        except Exception as e:
            raise Exception(f"Error reading data from Excel: {e}")

    def scenario_include_stable_sources(self):
        return any(source_name in self.sources_dict for source_name in
                   ['PPA','Grid', 'Gas Generator', 'HFO Generator', 'HFO+Gas Generator'])

    def get_gen_pwr_ops(self, source_type, unit_type, current_year):
        # Check if the given source_type exists in the sources dictionary
        if source_type not in self.sources_dict:
            raise ValueError(f"No source of type {source_type} found.")

        source = self.sources_dict[source_type]

        # Check if the unit_type is valid
        if unit_type not in ['PRIMARY', 'BACKUP']:
            raise ValueError(f"Invalid unit type {unit_type}.")

        # Determine which attributes to use based on the unit_type
        if unit_type == 'PRIMARY':
            count_key = 'count_prim_units'
            rating_key = 'rating_prim_units'
        else:  # 'BACKUP'
            count_key = 'count_backup_units'
            rating_key = 'rating_backup_units'
        perc_op_key = 'perc_rated_output'

        # If the source has the 'gas_fuel_type' attribute, then calculate capacity with derating
        if 'gas_fuel_type' in source.inputs:
            fuel_der_fac = self.derating_factor(source.inputs['gas_fuel_type'])
        else:
            fuel_der_fac = 1

        # Calculate total potential power with degradation
        total_pwr_pot = 0
        degradation_rate = source.meta.degradation if hasattr(source.meta, 'degradation') else 0
        total_count = 0
        for year, yr_data in source.inputs.items():
            if isinstance(year, int) and year <= current_year:
                years_of_operation = current_year - year
                if perc_op_key in yr_data:
                    perc_op = yr_data[perc_op_key]/100
                else:
                    perc_op = 1
                degradation_factor = 1 - (degradation_rate * years_of_operation/100)
                total_pwr_pot += yr_data[count_key] * yr_data[rating_key] * perc_op * \
                                 fuel_der_fac * degradation_factor
                total_count += yr_data[count_key]

                if year == current_year:
                    break

        return total_count, total_pwr_pot

    def get_gen_ener_op(self, source_type, current_year, current_month):

        if source_type not in self.sources_dict:
            raise ValueError(f"No source of type {source_type} found.")

        source = self.sources_dict[source_type]

        #NOT NEEDED IN ENERGY BECAUSE WE ALREADY ACCOUNT FOR THIS IN
        """
        if 'gas_fuel_type' in source.inputs:
            fuel_der_fac = self.derating_factor(source.inputs['gas_fuel_type'])
        else:
            fuel_der_fac = 1
        """

        # Retrieve the degradation rate
        degradation_rate = getattr(source.meta, 'degradation', 0)

        # Total Energy Potential considering straight-line degradation
        total_ener_pot = 0
        for year, yr_data in source.inputs.items():
            if isinstance(year, int) and year <= current_year:
                years_of_operation = current_year - year
                if  'perc_rated_output' in yr_data:
                    perc_op = yr_data['perc_rated_output']/100
                else:
                    perc_op = 1

                degradation_factor = 1 - (degradation_rate * years_of_operation/100)
                total_ener_pot += (yr_data['count_prim_units'] * yr_data['rating_prim_units'] *
                                   perc_op  * degradation_factor) * 24

                if year == current_year:
                    break

        # Multiply by the number of days in the current month to get the total energy potential for the month
        total_ener_pot *= self.ip_enr_data[current_month]['days']

        return total_ener_pot

    def _get_monthly_energy_req(self, year, month):

        prod_energy_req = self.ip_enr_data[month]['prod_enr_consump_per_MW_MD'] * \
                          self.ip_load_data[year]['max_dem_load_day']

        if self.ip_enr_data[month]['season'] == 'Winter':
            cool_dem_prop_ngt = self.ip_site_data['wint_nht_cool_dem_prop'] / 100
            cool_dem_prop_day = self.ip_site_data['wint_day_cool_dem_prop'] / 100
        else:
            cool_dem_prop_ngt = self.ip_site_data['summ_nht_cool_dem_prop'] / 100
            cool_dem_prop_day = self.ip_site_data['summ_nht_cool_dem_prop'] / 100

        cooling_energy_req = ((self.ip_enr_data[month]['length_night'] * cool_dem_prop_ngt *
                               self.ip_load_data[year]['cool_elect_load']) +
                              (self.ip_enr_data[month]['length_day'] * cool_dem_prop_day *
                               self.ip_load_data[year]['cool_elect_load'])) * self.ip_enr_data[month]['days']

        #Energy requirements to charge BESS
        if 'BESS' in self.sources_dict:

            #sum the failure per year of all sources

            tot_failures = sum(self.sources_dict[src_name].meta.num_failures_year
                               for src_name in self.available_sources(all_sources=False)
                               if src_name in self.sources_dict)
            charge_cycles_month = round(self.ip_enr_data[month]['days'] * (tot_failures / 365))

            _, total_cap = self.get_gen_pwr_ops('BESS', 'PRIMARY', year)

            BESS_charge_enr_req = charge_cycles_month * total_cap
        else:
            BESS_charge_enr_req = 0

        return prod_energy_req, cooling_energy_req, BESS_charge_enr_req

    def determine_pot_failures(self, src,year,month):

        # for Grid
        # monthly failures are independent of one another.
        if src.source_type == 'Grid' or 'PPA':
            monthly_failures = src.meta.num_failures_year / 12
            lower_bound = 0.75 * monthly_failures
            upper_bound = 1.25 * monthly_failures
            # Randomly round up or down for each bound
            lower_bound = math.ceil(lower_bound) if random.choice([True, False]) else math.floor(lower_bound)
            upper_bound = math.ceil(upper_bound) if random.choice([True, False]) else math.floor(upper_bound)

            # If lower_bound and upper_bound are equal, return one of them
            if lower_bound > upper_bound:
                return random.randint(upper_bound, lower_bound)
            elif lower_bound < upper_bound:
                return random.randint(lower_bound, upper_bound)
            else:
                return lower_bound
        else:
            # for all other sources
            num_failures_so_far = sum(src.outputs[year][m]['num_failures'] for m in range(1, month))
            poss_annual_failures = src.meta.num_failures_year
            remaining_failures = poss_annual_failures - num_failures_so_far
            months_left = 12 - month + 1

            # No more failures needed
            if remaining_failures <= 0:
                return 0

            # Expected failures this month
            expected_failures = remaining_failures / months_left

            # Randomly decide the number of failures this month
            monthly_failures = 0
            for _ in range(int(expected_failures * 2)):  # Adjust the multiplier for more randomness
                if random.random() < expected_failures / 2:  # Adjust the divisor for probability
                    monthly_failures += 1

        # Ensure the total failures don't exceed the annual limit

        return min(monthly_failures,remaining_failures)


    def calc_ins_backup_pwr_pot(self, year, month):
        wind_min_power = min([self.sources_dict['Wind'].calc_output_power(year, month, hour)
                              for hour in range(1, 25)]) if 'Wind' in self.sources_dict else 0
        bess_capacity = self.get_gen_pwr_ops('BESS', 'PRIMARY', year)[1] if 'BESS' in self.sources_dict else 0
        return wind_min_power + bess_capacity

    def grid_pk_to_offpk(self, month, energy):
        pk_to_nonpk_ratio = self.ip_enr_data[month]["grid_enr_pk_to_nonpk_prop"]
        nonpk_enr = energy / (1 + pk_to_nonpk_ratio)
        pk_enr = energy - nonpk_enr
        return pk_enr, nonpk_enr

    def free_cooling_enr_cal(self, year, month):
        total_cool_op = 0
        for src_name in ['Gas Generator', 'HFO Generator', 'HFO+Gas Generator']:

            if src_name in self.sources_dict:

                src = self.sources_dict[src_name]
                avg_src_pwr = sum([src.outputs[year][month][hour]['power_output_prim_units']
                                   for hour in range(1, 25)])/24
                src_cool_op = avg_src_pwr * src.meta.cooling_load_feeding_capability
                total_cool_op += src_cool_op
        cool_op_pwr = max(0, total_cool_op * 3.412 / (1000 * self.ip_site_data['cop_of_electric_chiller']))

        return cool_op_pwr * 24 * self.ip_enr_data[month]['days']

    def get_scenario_config(self):
        # List of all available sources
        #all_sources = self.available_sources()

        # Data structure to hold the extracted data
        data = {
            "Year": [i for i in range(0, self.n + 1)]
        }

        # For each year, gather the data for each source
        for year in range(0, self.n + 1):
            for src in self.sources_list:
                num_units = src.inputs[year]['count_prim_units']
                unit_rating = src.inputs[year]['rating_prim_units']
                total_capacity = num_units * unit_rating
                priority = src.priority

                # Append data to the structure
                #data.setdefault(f"{src} Num of Units", []).append(num_units)
                #data.setdefault(f"{src} Unit Rating MW", []).append(unit_rating)
                data.setdefault(f"{src.source_type} Total Capacity MW ", []).append(total_capacity)
                data.setdefault(f"{src.source_type} Source Priority", []).append(priority)
        # Convert the data structure to a DataFrame
        self.scenario_spec = pd.DataFrame(data)

    # CALCULATION FUNCTIONS
    def energy_calculation(self):

        for year in range(1, self.n + 1):
            for month in range(1, 13):
                month_data = {'year': year, 'month': month}
                print(f"Energy Calc Year {year}, month {month}")
                # Determine energy requirements
                prod_enr_req, cool_enr_req, bess_charge_enr_req = self._get_monthly_energy_req(year, month)

                month_data['Prod Energy Req, MWh'] = prod_enr_req
                month_data['Cooling Energy Req, MWh'] = cool_enr_req

                cool_enr_req = max(0,cool_enr_req - self.free_cooling_enr_cal(year, month))
                month_data['Cooling Energy Req after CHP adj., MWh'] = cool_enr_req
                month_data['BESS Charging Energy Req, MWh'] = bess_charge_enr_req

                month_tot_enr_req = prod_enr_req + cool_enr_req + bess_charge_enr_req
                month_data['Total Energy Req, MWh'] = month_tot_enr_req
                month_rem_enr_req = month_tot_enr_req

                critical_load = self.ip_load_data[year]['crit_load_prop'] * self.ip_load_data[year]['max_dem_load_day'] / 100

                # Energy from renewables
                for ren_src_name in ['Wind', 'Solar','PPA']:

                    if ren_src_name in self.sources_dict:
                        print(f"Finding {ren_src_name} energy")
                        pot_enr_op = self.sources_dict[ren_src_name].calc_output_energy(year, month)

                        # Wind energy func returns daily energy value
                        if ren_src_name == 'Wind':
                            pot_enr_op *= self.ip_enr_data[month]['days']
                        ren_enr_op = min(month_rem_enr_req, pot_enr_op)
                        month_rem_enr_req -= ren_enr_op
                        self.sources_dict[ren_src_name].outputs[year][month]['energy_output_prim_units'] = ren_enr_op
                        month_data[f"{ren_src_name} Output in MWh"] = ren_enr_op
                month_data["Remaining Energy Demand (after Renewables) MWh"] = month_rem_enr_req
                 
                for src in self.sources_list:

                    if src.source_type in self.stable_sources(include_backup=False):
                        src_name = src.source_type
                        print(f"Finding {src_name} potential energy")
                        gen_pot_enr_op = self.get_gen_ener_op(src_name, year, month) - \
                                         src.outputs[year][month]['energy_output_prim_units']

                        # Calculate Monthly Failure Probability
                        print(f"Finding failures for {src_name}")
                        num_pot_failures = self.determine_pot_failures(src,year,month)
                        month_data[f'{src_name} Potential Failures'] = num_pot_failures
                        if num_pot_failures == 0:
                            month_data[f'{src_name} Failures mitigated'] = 0
                            month_data[f'{src_name} Unavailability, hrs'] = 0
                        else:
                            # determined reduced energyfind energy required to cover each failure
                            month_data[f'{src_name} Potential Failures'] = num_pot_failures
                            hourly_pot = gen_pot_enr_op/(self.ip_enr_data[month]['days']*24)
                            lost_enr_pot = num_pot_failures * hourly_pot * src.meta.avg_failure_time
                            gen_pot_enr_op -= lost_enr_pot
                            en_per_fail = src.meta.avg_failure_time * critical_load
                            num_fails_not_cov = num_pot_failures

                            #look through other primary sources
                            for alt_src in self.sources_list:
                                if alt_src.source_type in self.stable_sources(include_backup=True) \
                                        and alt_src.source_type != src.source_type:

                                    alt_src_name = alt_src.source_type
                                    print(f"Checking if {alt_src_name} can provide failure coverage {src_name}")
                                    _, total_cap = self.get_gen_pwr_ops(alt_src_name, 'PRIMARY', year)
                                    # ...and check if these can kick in to cover the failure duration
                                    if total_cap >= critical_load:

                                        print(f"{alt_src_name} does have power cap to backup {src_name}")
                                        # how many failures can be alternate source cover in terms of energy
                                        alt_src_en_pot = self.get_gen_ener_op(alt_src_name, year, month)
                                        alt_src_en_rem = alt_src_en_pot - \
                                                         alt_src.outputs[year][month]['energy_output_prim_units']

                                        alt_src_nfail_cover = math.floor(alt_src_en_rem / en_per_fail)
                                        if not alt_src_nfail_cover:
                                            alt_src_nfail_cover = 0
                                        alt_src_nfail_cover = min(alt_src_nfail_cover, num_fails_not_cov)

                                        # add the failure coverage energy to the alt source's expenditure
                                        # remaining failures are reduced and other sources may cover them (loop)
                                        if alt_src.source_type == 'Grid':
                                            backup_enr_pk, backup_enr_nonpk = \
                                                self.grid_pk_to_offpk(month, alt_src_nfail_cover * en_per_fail)
                                            alt_src.outputs[year][month]['energy_output_peak'] += backup_enr_pk
                                            alt_src.outputs[year][month]['energy_output_offpeak'] += backup_enr_nonpk

                                        elif alt_src.source_type == 'HFO+Gas Generator':
                                            gas_enr, hfo_enr = alt_src.gas_hfo_enr_op(alt_src_nfail_cover * en_per_fail)
                                            alt_src.outputs[year][month]['energy_output_prim_units'] += gas_enr
                                            alt_src.outputs[year][month]['energy_output_prim_units_sec'] += hfo_enr

                                        else:
                                            alt_src.outputs[year][month]['energy_output_prim_units'] += \
                                                (alt_src_nfail_cover * en_per_fail)
                                        num_fails_not_cov -= alt_src_nfail_cover

                                        # if potential failures have been reduced to zero
                                        # then further sources don't need to tried.
                                        if num_fails_not_cov <= 0:
                                            break
                            # Calculate Instant Backup Potential Power
                            ins_backup_pot_pwr = self.calc_ins_backup_pwr_pot(year, month)

                            if ins_backup_pot_pwr >= critical_load:
                                num_failures = num_fails_not_cov
                            else:
                                num_failures = num_pot_failures
                            failure_duration = num_fails_not_cov * src.meta.avg_failure_time

                            month_data[f'{src_name} Failures mitigated'] = num_pot_failures - num_failures
                            month_data[f'{src_name} Unavailability, hrs'] = failure_duration

                            src.outputs[year][month]['num_pot_failures'] = num_pot_failures
                            src.outputs[year][month]['num_failures'] = num_failures
                            src.outputs[year][month]['failure_duration'] = failure_duration

                        # Energy output calculation for stable sources (including failure adjustments)
                        print(f"Finding the energy output for {src_name}")

                        gen_enr_op = min(month_rem_enr_req, gen_pot_enr_op)
                        month_rem_enr_req -= gen_enr_op
                        if src.source_type == 'Grid':
                            enr_pk, enr_nonpk = self.grid_pk_to_offpk(month, gen_enr_op)
                            src.outputs[year][month]['energy_output_peak'] += enr_pk
                            src.outputs[year][month]['energy_output_offpeak'] += enr_nonpk
                            month_data['Grid Peak Energy, MWh'] = src.outputs[year][month]['energy_output_peak']
                            month_data['Grid Off Peak Energy, MWh'] = src.outputs[year][month]['energy_output_offpeak']

                        elif src.source_type == 'HFO+Gas Generator':
                            gas_enr, hfo_enr = src.gas_hfo_enr_op(gen_enr_op)
                            src.outputs[year][month]['energy_output_prim_units'] += gas_enr
                            src.outputs[year][month]['energy_output_prim_units_sec'] += hfo_enr
                            month_data['HFO+Gas Gen, Energy from HFO, MWh'] = src.outputs[year][month]['energy_output_prim_units_sec']
                            month_data['HFO+Gas Gen, Energy from Gas, MWh'] = src.outputs[year][month]['energy_output_prim_units']

                        else:
                            src.outputs[year][month]['energy_output_prim_units'] += gen_enr_op
                            month_data[f"{src_name} Output in MWh"] = \
                                self.sources_dict[src_name].outputs[year][month]['energy_output_prim_units']

                        #if month_rem_enr_req <= 0:
                            #break
                if 'Diesel Generator' in self.sources_dict:
                    src_name = 'Diesel Generator'
                    if month_rem_enr_req > 0: #if 'Diesel Generator' in self.sources_dict and month_rem_enr_req > 0:
                        print("Finding the energy output for Diesel Generator")
                        
                        gen_pot_enr_op = self.get_gen_ener_op(src_name, year, month) - \
                                        src.outputs[year][month]['energy_output_prim_units']
                        gen_enr_op = min(month_rem_enr_req, gen_pot_enr_op)
                        month_rem_enr_req -= gen_enr_op
                    else:
                        gen_enr_op = 0
                    self.sources_dict[src_name].outputs[year][month]['energy_output_prim_units'] += gen_enr_op
                    month_data[f"{src_name} Output in MWh"] = \
                        self.sources_dict[src_name].outputs[year][month]['energy_output_prim_units']

                month_data['Final Unserved Energy Req in MWh'] = month_rem_enr_req
                self.energy_df.append(month_data)
                print(f"Energy data for the year {year}, month {month} determined. Unserved is {month_rem_enr_req} MWh")
        self.energy_df = pd.DataFrame(self.energy_df)

    def power_calculation2(self):

        for year in range(1, self.n + 1):
            for month in range(1, 13):
                for hour in range(1, 25):
                    hour_data = {'year': year, 'month': month, 'hour': hour}

                    # Determine Production Demand
                    if hour > 19 or hour < 6:
                        hour_data['Prod Demand in MW'] = self.ip_load_data[year]['max_dem_load_night']
                    else:
                        hour_data['Prod Demand in MW'] = self.ip_load_data[year]['max_dem_load_day']

                    prod_demand = hour_data['Prod Demand in MW']

                    hour_data['Cooling demand in TR'] = self.ip_load_data[year]['cool_req_in_tr']

                    free_cooling_output = 0
                    cooling_elect_load = 0

                    # Calculate Free Cooling
                    for gen in ['Gas Generator', 'Existing Gas Generators','HFO Generator', 'HFO+Gas Generator']:
                        if gen in self.sources_dict:
                            _, total_cap = self.get_gen_pwr_ops(gen, 'PRIMARY', year)
                            free_cooling_output += total_cap * self.sources_dict[gen].meta.cooling_load_feeding_capability

                    hour_data['Free Cooling available in TR'] = free_cooling_output
                    rem_cooling_demand = hour_data['Cooling demand in TR'] - free_cooling_output
                    cooling_elect_load = max(0, rem_cooling_demand * 3.412 / (
                            1000 * self.ip_site_data['cop_of_electric_chiller']))
                    #hour_data['Total power demand incl. Cooling in MW'] = \
                    #    hour_data['Prod Demand in MW'] + cooling_elect_load

                    bess_charge_load = 0
                    # BESS charging impact
                    if 'BESS' in self.sources_dict:
                        _, total_cap = self.get_gen_pwr_ops('BESS', 'PRIMARY', year)
                        hour_data['BESS Charging Demand'] = total_cap * 0.25
                        bess_charge_load = hour_data['BESS Charging Demand']

                    unserved_demand = bess_charge_load + cooling_elect_load + prod_demand
                    hour_data['Total power demand incl. Cooling & BESS in MW'] = unserved_demand

                    # Satisfy Demand with Sources
                    for src in self.sources_list:
                        
                        src_name = src.source_type

                        if src_name in ['Solar', 'Wind','PPA']:
                            output_potential = self.sources_dict[src_name].calc_output_power(year, month, hour)
                        elif src_name == "BESS":
                            continue
                        else:
                            _, output_potential = self.get_gen_pwr_ops(src_name, 'PRIMARY', year)

                        output_actual = min(unserved_demand, output_potential)
                        unserved_demand -= output_actual
                        unserved_demand = max(0, unserved_demand)
                        hour_data[f'{src_name} Output in MW'] = output_actual
                        hour_data[f'{src_name} Loading in %'] = (output_actual * 100) / output_potential if output_potential else 0

                        self.sources_dict[src_name].outputs[year][month][hour]['power_output_prim_units'] = output_actual
                        self.sources_dict[src_name].outputs[year][month][hour]['loading_prim_units'] = hour_data[f'{src_name} Loading in %']

                    hour_data['Final Unserved Load'] = unserved_demand

                    # Append the hour's data to the list
                    self.power_df.append(hour_data)
        self.power_df = pd.DataFrame(self.power_df)

    def opex_calculation(self):
        fuel_tariff = FuelTariff()  # Create an instance of the FuelTariff class
        for y in range(1, self.n + 1):

            for m in range(1, 13):

                month_data = {'year': y, 'month': m}
                interrupt_loss = 0
                outage_loss = 0
                interrupt_list = []
                outage_list = []
                for src in self.sources_list:

                    src_name = src.source_type
                    # Calculate total capacity up till current year
                    total_capacity = sum(
                        yr_data['count_prim_units'] * yr_data['rating_prim_units']
                        for yr, yr_data in src.inputs.items() if isinstance(yr, int) and yr <= y
                    )
                    if src_name == "Existing Gas Generators":
                        if total_capacity > 0:
                            src.outputs[y]['depreciation_cost'] = src.meta.existing_cap_cost / src.meta.useful_life
                        else:
                            src.outputs[y]['depreciation_cost'] = 0
                    
                    elif src_name != 'Grid' and src_name != 'PPA':

                        total_capex = sum([src.outputs[y]['capital_cost'] for y in range(y + 1)])
                        # Compute the annual depreciation
                        src.outputs[y]['depreciation_cost'] = total_capex / src.meta.useful_life

                    inflation_rate = pow(1 + src.meta.opex_inflation_rate, y)
                    src_mnth_op = src.outputs[y][m]

                    # Calculate and save fixed OPEX
                    if src_name != 'Grid' and src_name != 'PPA':
                        src_mnth_op['fixed_opex'] = total_capacity * \
                                                    src.meta.fixed_opex_baseline * inflation_rate

                        month_data[f'{src_name} Depreciation Cost, M PKR'] = \
                            src.outputs[y]['depreciation_cost'] / (12 * 1000000)
                        month_data[f'{src_name} Fixed Opex, M PKR'] = src_mnth_op['fixed_opex'] / 1000000
                    else:
                        src_mnth_op['fixed_charges'] = total_capacity * \
                                                       src.meta.tariff_baseline_fixed * inflation_rate
                        month_data[f'{src_name} Fixed Opex, M PKR'] = src_mnth_op['fixed_charges'] / 1000000

                    # Calculate and save variable OPEX
                    if src_name not in ['BESS', 'Solar', 'Wind', 'PPA','Grid']:
                        src_mnth_op['var_opex'] = src_mnth_op['energy_output_prim_units'] * \
                                                  src.meta.var_opex_baseline * inflation_rate
                        month_data[f'{src_name} Var OPEX, M PKR'] = src_mnth_op['var_opex'] / 1000000

                    # Calculate and save energy charges for Grid
                    if src_name == 'Grid':
                        src_mnth_op['peak_enr_charges'] = src_mnth_op['energy_output_peak'] * \
                                                          src.meta.tariff_baseline_var_peak * inflation_rate
                        src_mnth_op['offpeak_enr_charges'] = src_mnth_op['energy_output_offpeak'] * \
                                                             src.meta.tariff_baseline_var_offpeak * inflation_rate
                        month_data['Grid Peak Rate Energy Cost, M PKR'] = src_mnth_op['peak_enr_charges'] / 1000000
                        month_data['Grid Offpeak Rate Energy Cost, M PKR'] = src_mnth_op['offpeak_enr_charges'] / 1000000
                    
                    if src_name == 'PPA':
                        src_mnth_op['enr_charges'] = src_mnth_op['energy_output_prim_units'] * \
                                                          src.meta.tariff_baseline_var * inflation_rate
                        month_data['PPA Energy Cost, M PKR'] = src_mnth_op['enr_charges'] / 1000000

                    # Calculate and save fuel costs
                    if src_name in ['Gas Generator', 'Existing Gas Generators','HFO Generator', 'HFO+Gas Generator', 'Diesel Generator']:

                        fuel_type = src.inputs['fuel_type']
                        fuel_data = fuel_tariff.get_tariff_and_inflation(fuel_type)

                        # Initialize fuel charges for the month
                        src_mnth_op['fuel_charges'] = 0

                        for year in src.inputs:
                            if isinstance(year, int) and year <= y:
                                y_capacity = src.inputs[year]['count_prim_units'] * src.inputs[year]['rating_prim_units']
                                proportion = y_capacity / total_capacity if total_capacity else 0
                                energy_output_component = src_mnth_op['energy_output_prim_units'] * proportion
                                fuel_eff_factor = 1 + (100 - src.inputs[year]['fuel_eff']) / 100
                                # Calculate fuel charges for the energy output component of year y
                                src_mnth_op['fuel_charges'] += energy_output_component * fuel_data['tariff'] * fuel_eff_factor * pow(1 + fuel_data['inflation'], y)

                        # Secondary fuel charges for HFO+Gas Generator
                        if src_name == 'HFO+Gas Generator':
                            sec_fuel_type = src.inputs['sec_fuel_type']
                            fuel_data_sec = fuel_tariff.get_tariff_and_inflation(sec_fuel_type)
                            src_mnth_op['fuel_charges_sec'] = 0  # Initialize secondary fuel charges for the month

                            # Calculate secondary fuel charges for each year's additions up till the current year (year)
                            for year in src.inputs:
                                if isinstance(year, int) and year <= y:
                                    y_capacity = src.inputs[year]['count_prim_units'] * src.inputs[year]['rating_prim_units']
                                    proportion = y_capacity / total_capacity if total_capacity else 0
                                    energy_output_component_sec = src_mnth_op['energy_output_prim_units_sec'] * proportion
                                    fuel_eff_factor_sec = 1 + (100 - src.inputs[year]['fuel_eff']) / 100
                                    # Calculate secondary fuel charges for the energy output component of year y
                                    src_mnth_op['fuel_charges_sec'] += energy_output_component_sec * fuel_data_sec['tariff'] * fuel_eff_factor_sec * pow(1 + fuel_data_sec['inflation'], y)

                            # Convert total secondary fuel charges to M PKR
                            month_data[f'Fuel Charges for {sec_fuel_type}, M PKR'] = src_mnth_op['fuel_charges_sec'] / 1000000

                        # Convert total primary fuel charges to M PKR
                        month_data[f'Fuel Charges for {fuel_type}, M PKR'] = src_mnth_op['fuel_charges'] / 1000000

                    # Determine number of interruptions
                    interrupt_list.append(src.outputs[y][m]['num_failures'])
                    outage_list.append(src.outputs[y][m]['failure_duration'])

                # since not all interruptions would impact critical load
                # we take the average of interruptions (not mitigated
                month_data['Loss due to Interruptions, M PKR'] = round(sum(interrupt_list)/len(interrupt_list)) * \
                                                                 self.ip_site_data['fail_loss_immediate'] / 1000000
                month_data['Loss due to Outage, M PKR'] = round(sum(outage_list)/len(outage_list)) * \
                                                          self.ip_site_data['fail_loss_over_time'] / 1000000
                self.opex_df.append(month_data)
        self.opex_df = pd.DataFrame(self.opex_df)

    def emissions_calculation(self):

        fuel_struct = FuelTariff()

        for y in range(1, self.n + 1):
            for m in range(1, 13):
                month_data = {'year': y, 'month': m}

                for src in self.sources_list:
                    if src.source_type in self.stable_sources(include_backup=True):
                        src_name = src.source_type

                        if src_name == 'Grid':
                            src.outputs[y][m]['co2_emissions'] = \
                                (src.outputs[y][m]['energy_output_peak'] +
                                src.outputs[y][m]['energy_output_offpeak']) * src.meta.co2_emission
                        
                        elif src_name == 'HFO+Gas Generator':
                            fuel_data = fuel_struct.get_tariff_and_inflation(src.inputs['sec_fuel_type'])
                            src.outputs[y][m]['co2_emissions'] += src.outputs[y][m]['energy_output_prim_units_sec'] \
                                                                * fuel_data['co2_emission']            
                        else:
                            fuel_data = fuel_struct.get_tariff_and_inflation(src.inputs['fuel_type'])
                            src.outputs[y][m]['co2_emissions'] = src.outputs[y][m]['energy_output_prim_units'] \
                                                                * fuel_data['co2_emission']
                        month_data[f'CO2 Emissions from {src_name}, MT'] = src.outputs[y][m]['co2_emissions']/1000
                self.emissions_df.append(month_data)
        self.emissions_df = pd.DataFrame(self.emissions_df)

    def capex_calculation2(self):

        for year in range(0, self.n + 1):
            y_data = {'year': year}
            for src in self.sources_list:

                src_name = src.source_type
                
                # Get the required values
                count_prim_units = src.inputs[year]['count_prim_units']
                rating_prim_units = src.inputs[year]['rating_prim_units']
                cap_cost_baseline = src.meta.capital_cost_baseline
                cap_cost_y_zero = src.meta.existing_cap_cost
                capex = 0
                if src_name == "Existing Gas Generators":
                    if count_prim_units != 0:
                        capex = src.meta.ingestion_cost * pow(1 + self.ip_site_data['capital_inflation_rate'], year)
                else:

                    # Calculate the capital cost for this year and source
                    if year == 0:
                        if count_prim_units == 0:
                            capex = 0
                        else:
                            capex = cap_cost_y_zero
                    else:
                        capex = (count_prim_units * rating_prim_units * cap_cost_baseline
                                * pow(1 + self.ip_site_data['capital_inflation_rate'], year))

                    # Store the calculated capex in the source's outputs structure
                src.outputs[year]['capital_cost'] = int(capex)
                y_data[f'{src_name} CAPEX, M PKR'] = int(capex / 1000000)
            self.capex_df.append(y_data)
        self.capex_df = pd.DataFrame(self.capex_df)

    #SUMMARY FUNCTIONS
    def gen_annual_summary(self):

        # Power and Energy Fulfillment Factors for each year
        power_fulfillment_per_year = self.power_df.groupby('year').apply(
            lambda df: (len(df[df['Final Unserved Load'] == 0]) / len(df)) * 100).reset_index(
            name='Power Fulfilment Factor, %')
        energy_fulfillment_per_year = self.energy_df.groupby('year').apply(
            lambda df: (len(df[df['Final Unserved Energy Req in MWh'] == 0]) / len(df)) * 100).reset_index(
            name='Energy Fulfillment Factor, %')

        # Total Annual Energy Production
        annual_energy_production = self.energy_df.groupby('year').apply(
            lambda df: (df['Total Energy Req, MWh'] - df['Final Unserved Energy Req in MWh']).sum()).reset_index(
            name='Energy Production, MWh')

        # Annual CAPEX, OPEX, and emissions
        annual_opex = self.opex_df.drop(columns=['year', 'month']).groupby(self.opex_df['year']).sum().sum(
            axis=1).reset_index(name='OPEX, M PKR')
        annual_emissions = self.emissions_df.drop(columns=['year', 'month']).groupby(self.emissions_df['year']).sum().sum(
            axis=1).reset_index(name='CO2 Emissions, MT')

        # Merge all dataframes together and calculate Equivalent Tariff
        summary_df = power_fulfillment_per_year.merge(energy_fulfillment_per_year, on='year') \
            .merge(annual_energy_production, on='year'). \
            merge(annual_opex, on='year'). \
            merge(annual_emissions, on='year')

        # Calculate Equivalent Tariff
        summary_df['Equivalent Tariff, PKR/kWh'] = summary_df['OPEX, M PKR'] * 1000/ summary_df['Energy Production, MWh']
        self.summary_df = summary_df

    def gen_enr_fulfillment(self):
        # Extract rows where 'Final Unserved Energy Req in MWh' is greater than zero
        self.unserved_energy = self.energy_df[self.energy_df['Final Unserved Energy Req in MWh'] > 0]

        # Calculate the 'Energy Fulfillment Factor'
        total_rows = len(self.energy_df)
        served_energy_rows_count = len(self.energy_df[self.energy_df['Final Unserved Energy Req in MWh'] == 0])
        self.energy_fulfillment = (served_energy_rows_count / total_rows) * 100

    def gen_pwr_fulfillment(self):

        self.unserved_power = self.power_df[self.power_df['Final Unserved Load'] > 0]

        # Calculate the 'Power Fulfillment Factor'
        total_power_rows = len(self.power_df)
        served_power_rows_count = len(self.power_df[self.power_df['Final Unserved Load'] == 0])
        self.power_fulfillment = (served_power_rows_count / total_power_rows) * 100

    def gen_annual_pwr_summary(self):

        self.power_summary_df = self.power_df.groupby('year').apply(self.select_row_for_year).reset_index(drop=True).drop(columns=['Difference'])

    def gen_annual_enr_summary(self):
        self.energy_summary_df = self.energy_df.groupby('year').sum().reset_index()
        self.energy_summary_df.drop(columns=['month'], inplace=True)

    def gen_annual_opex_summary(self):
        self.opex_summary_df = self.opex_df.groupby('year').sum().reset_index()
        self.opex_summary_df.drop(columns=['month'], inplace=True)

    def gen_annual_emissions_summary(self):
        self.emissions_summary_df = self.emissions_df.groupby('year').sum().reset_index()
        self.emissions_summary_df.drop(columns=['month'], inplace=True)

    def gen_annual_enr_summary_concise(self):

        self.energy_summary_concise_df = self.energy_df.groupby('year').sum().reset_index()
        self.energy_summary_concise_df.drop(columns=['month'], inplace=True)

        cols_to_drop = self.energy_summary_concise_df.columns[self.energy_summary_concise_df.columns.str.contains(
            'Failures|Unavailability|Cooling|BESS|Prod Energy|Remaining')]
        self.energy_summary_concise_df.drop(columns=cols_to_drop, inplace=True)

    def gen_annual_opex_summary_concise(self):

        # 1. Group by year
        grouped = self.opex_df.groupby('year').sum().reset_index()

        # 2. Drop the month column.
        grouped.drop('month', axis=1, inplace=True)

        # 3. Sum all column values that contain 'Depreciation'
        depreciation_cols = grouped.filter(like='Depreciation').columns
        grouped['Source Depreciation, M PKR'] = grouped[depreciation_cols].sum(axis=1)
        grouped.drop(depreciation_cols, axis=1, inplace=True)

        # 4. Sum columns that contain either 'Solar' or 'Wind'
        renewable_cols = grouped.filter(like='Solar').columns.tolist() + grouped.filter(like='Wind').columns.tolist()
        grouped['Renewable OPEX, M PKR'] = grouped[renewable_cols].sum(axis=1)
        grouped.drop(renewable_cols, axis=1, inplace=True)

        # 5. Sum all columns that contain the word 'Generator' and 'OPEX' (except Diesel Generator)
        power_plant_cols = [col for col in grouped.columns if
                            'Generator' in col and 'OPEX' in col and 'Diesel' not in col]
        grouped['Power Plant OPEX, M PKR'] = grouped[power_plant_cols].sum(axis=1)
        grouped.drop(power_plant_cols, axis=1, inplace=True)

        # 6. Sum all columns that contain Diesel (exclude Diesel Generator Depreciation Cost which is already dropped)
        diesel_cols = grouped.filter(like='Diesel').columns
        grouped['Diesel OPEX, M PKR'] = grouped[diesel_cols].sum(axis=1)
        grouped.drop(diesel_cols, axis=1, inplace=True)

        # 7. Sum all columns that contain Grid
        grid_cols = grouped.filter(like='Grid').columns
        grouped['Grid OPEX, M PKR'] = grouped[grid_cols].sum(axis=1)
        grouped.drop(grid_cols, axis=1, inplace=True)

        # 8. Sum all columns that contain PPA
        ppa_cols = grouped.filter(like='PPA').columns
        grouped['PPA OPEX, M PKR'] = grouped[ppa_cols].sum(axis=1)
        grouped.drop(ppa_cols, axis=1, inplace=True)

        self.opex_summary_concise_df = grouped

    def select_row_for_year(self, group):
        group['Difference'] = (
                group['Total power demand incl. Cooling & BESS in MW'] - group['Final Unserved Load']).abs()
        min_difference = group['Difference'].min()
        selected_rows = group[group['Difference'] == min_difference]
        return selected_rows.sample(1)  # If there are multiple rows with the same minimum difference, pick a random one

